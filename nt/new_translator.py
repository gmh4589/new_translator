# -*- coding:utf-8 -*-
from kivy.app import App
from kivy.factory import Factory
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from tkinter import filedialog as fd

import openpyxl

sss = 0

# Загружает конфиг интерфейса
Builder.load_file('new_translator.kv')

class MainScreen(Screen):

    values = []
    filename = ''

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    def fileOpen(self):

        filetypes = (
            ('Файлы Excel', '*.xlsx'),
            ('All files', '*.*')
        )

        self.filename = fd.askopenfilename(
            title = 'Выберите XLSX файл',
            initialdir = '/',
            filetypes = filetypes)

        print(self.filename)

        if self.filename != '':
            rearData = openpyxl.load_workbook(self.filename, data_only = True)
            sheet = rearData.active
            rows = sheet.max_row
            cols = sheet.max_column

            for row in range(1, rows):
                for col in range(1, cols):
                    cell = sheet.cell(row = row, column = col).value
                    if cell == 'dialog':
                        val = [sheet.cell(row = row + 1, column = col).value, '', row + 1, col]
                        self.values.append(val)

            print(self.values)

            self.ids['originalTXT'].text = str(self.values[0][0])
            self.ids['newTXT'].text = str(self.values[0][0])
            self.ids['rdyLabel'].text = '1/' + str(len(self.values) - 1) + ' готово'
            self.ids['longLabel'].text = str(len(self.values[0][1])) + ' символов'
            self.ids['percentLabel'].text = str((100 / len(self.values) * sss + 1))[:4] + ' %'

    def nextString(self, step = '+'):

        global sss
        global values

        try:
            if step == '+':
                if self.ids['originalTXT'].text != self.ids['newTXT'].text:
                    self.values[sss][1] = self.ids['newTXT'].text
                sss += 1
            elif step == '-' :
                if sss != -1:
                    if self.ids['originalTXT'].text == self.ids['newTXT'].text:
                        self.values[sss][1] = self.ids['newTXT'].text
                else:
                    Factory.EOFPopup().open()
                    sss = 0
                sss -= 1

            print(self.values)

            self.ids['originalTXT'].text = str(self.values[sss][0])
            if str(self.values[sss][1]) == '': self.ids['newTXT'].text = str(self.values[sss][0])
            else: self.ids['newTXT'].text = str(self.values[sss][1])
            self.ids['rdyLabel'].text = str(sss + 1) + '/' + str(len(self.values)) + ' готово'
            self.ids['longLabel'].text = str(len(self.values[sss][1])) + ' символов'
            self.ids['percentLabel'].text = str((100 / len(self.values) * sss))[:4] + ' %'
        except IndexError: Factory.EOFPopup().open()

    def reFresh(self):
        self.ids['newTXT'].text = self.ids['originalTXT'].text



class NewTranslatorApp(App):

    # Создаёт  интерфейс
    def build(self):
        sm = ScreenManager()
        sm.add_widget(MainScreen(name = 'main'))

        return sm

NewTranslatorApp().run()
