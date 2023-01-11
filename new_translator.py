# -*- coding:utf-8 -*-
import os
import random
import re

from kivy.app import App
from kivy.factory import Factory
from kivy.lang import Builder
from kivy.clock import Clock
from kivy.uix.screenmanager import ScreenManager, Screen

from tkinter import filedialog as fd
from tkinter import messagebox as mb

import openpyxl

sss = 0

# Загружает конфиг интерфейса
Builder.load_file('new_translator.kv')


class MainScreen(Screen):
    values = []
    filename = ''
    sheet = ''
    readData = ''
    sss = 0

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        Clock.schedule_interval(self.update, .1)
        Clock.schedule_interval(self.autoSave, 60)

    def fileOpen(self):

        if self.filename == '':
            filetypes = (
                ('Файлы Excel', '*.xlsx'),
                ('All files', '*.*')
            )

            self.filename = fd.askopenfilename(
                title = 'Выберите XLSX файл',
                initialdir = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') ,
                filetypes = filetypes)

            print(self.filename)

            if self.filename != '':
                self.readData = openpyxl.load_workbook(self.filename, data_only = True)
                self.sheet = self.readData.active
                rows = self.sheet.max_row
                cols = self.sheet.max_column

                for row in range(1, rows):
                    for col in range(1, cols):
                        cell = self.sheet.cell(row = row, column = col).value

                        if cell == 'dialog':
                            val = [self.sheet.cell(row = row + 1, column = col).value, '', row + 1, col]
                            self.values.append(val)

                self.upd()

                print(self.values)

        else: mb.showinfo(title = 'Сообщение', message = 'Файл уже открыт!')

    def writeFile(self, auto = 0):

        for cell in range(len(self.values)):
            o = 1 if self.values[cell][1] != '' else 0
            self.sheet.cell(row = int(self.values[cell][2]),
                            column = int(self.values[cell][3])).value \
                = self.values[cell][o]

        self.readData.save(self.filename)
        if auto == 0:
            Factory.SavedPopup().open()
        else:
            print('Автосохренение выполнено!')

    def autoSave(self, dt):
        self.writeFile(1)

    def nextString(self, step = '+'):

        global sss

        def setString():
            self.values[sss][1] = self.ids['newTXT'].text
            print(self.values[sss][1])

        try:
            setString()
            if step == '+':
                sss += 1
            elif step == '-':
                if sss != 0:
                    sss -= 1
                else:
                    Factory.EOFPopup().open()
            else:
                sss = step

            o = 1 if self.values[sss][1] != '' else 0
            self.ids['newTXT'].text = str(self.values[sss][o])

        except IndexError:
            Factory.EOFPopup().open()
        print(self.values)

    def update(self, dt):
        global load
        try:
            self.ids['originalTXT'].text = str(self.values[sss][0])
            self.ids['rdyLabel'].text = str(sss + 1) + '/' + str(len(self.values)) + ' готово'
            self.ids['longLabel'].text = str(len(self.values[sss][0])) + ' символов'
            self.ids['percentLabel'].text = str((100 / len(self.values) * sss))[:4] + ' %'

        except IndexError:pass

    def reFresh(self):
        self.ids['newTXT'].text = self.ids['originalTXT'].text

    # noinspection PyGlobalUndefined
    def findBTN(self):
        global count
        text4find = str(self.ids['text4find'].text)
        text4replace = str(self.ids['text4replace'].text)
        count = 0

        for st in range(len(self.values)):
            o = 1 if self.values[st][1] != '' else 0
            if self.values[st][o].find(text4find) != -1:
                self.values[st][1] = self.values[st][o].replace(text4find, text4replace)
                if st == sss: self.upd()
                count += 1

        mb.showinfo(title = 'Сообщение', message = f'Выполнено {str(count)} замен')
        print(self.values)

    def upd(self):
        try:
            if str(self.values[sss][1]) == '':
                self.ids['newTXT'].text = str(self.values[sss][0])
            else:
                self.ids['newTXT'].text = str(self.values[sss][1])
        except IndexError:
            mb.showinfo(title = 'Сообщение', message = f'В выбранном файле нет текста!')

    def gotoBTN(self):
        try:
            goto = self.ids['gotoInput'].text
            if int(goto) < len(self.values) + 1:
                self.nextString(goto - 1)
            else:
                mb.showinfo(title = 'Сообщение', message = f'Вы ввели {goto}, а в файле только {len(self.values)} строк!')
        except (TypeError, ValueError):
            mb.showinfo(title = 'Сообщение', message = f'Введите число от 1 до {len(self.values)}!')


class NewTranslatorApp(App):
    dicFile = open('./data/dictionary/ru.txt', 'r+', encoding = 'utf-8')
    dictionary = dicFile.readlines()
    # example = dictionary[random.randint(0, len(dictionary))]
    # print(example)
    # mb.showinfo(title = 'Сообщение', message = example)

    # Создаёт  интерфейс
    def build(self):
        sm = ScreenManager()
        sm.add_widget(MainScreen(name = 'main'))

        return sm


NewTranslatorApp().run()
